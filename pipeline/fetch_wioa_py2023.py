"""Fetch WIOA state performance, DOL ETA, for PY2023 explicitly.

The live "Results At-A-Glance" page now serves PY2024, so this fetcher does
NOT scrape it. It pulls the PY2023 National Performance Summary (PDF) and the
PY2023 Annual Report Accessible File (Excel) directly. State tables come from
the Excel (openpyxl); the national headline summary can be cross-read from the
PDF when a PDF text backend is available.

The fetcher takes a program_year parameter so other program years can be
backfilled later by swapping the year in the file URLs.

Title I metrics captured per jurisdiction: Adult employment rate Q2 after
exit, Adult median earnings Q2, Adult credential attainment, plus the Q4 and
measurable-skill-gains fields when present. No scoring; raw values only.
"""

from __future__ import annotations

import os

import states
from http_util import FetchBlocked, fetch, utc_now_iso
from manual import find_manual
from model import SourceFetch

SOURCE_ID = "wioa"
AT_A_GLANCE = "https://www.dol.gov/agencies/eta/performance/wioa-performance"
ARCHIVE = "https://www.dol.gov/agencies/eta/performance/results-archive"


def _pdf_backend_available() -> bool:
    """Lazily probe for a usable PDF text backend without polluting output.

    pypdf and pdfminer import the cryptography Rust binding at load time; where
    that binding is missing the import raises a pyo3 PanicException (a
    BaseException) and the Rust panic hook writes to OS-level stderr (fd 2),
    which a Python-level redirect cannot capture. We redirect fd 2 to devnull
    for the duration of the probe, and only run it when a PDF was actually
    downloaded, so a normal blocked run never triggers it. WIOA state and
    national figures come from the Excel; the PDF is a redundant, archival
    cross-check.
    """
    saved_fd = os.dup(2)
    devnull = os.open(os.devnull, os.O_WRONLY)
    try:
        os.dup2(devnull, 2)
        try:
            import pypdf  # noqa: F401
            return True
        except BaseException:  # noqa: BLE001
            return False
    finally:
        os.dup2(saved_fd, 2)
        os.close(devnull)
        os.close(saved_fd)


def _file_urls(program_year: int) -> dict[str, str]:
    """PY-specific PDF and Excel URLs. Backfill swaps the year token."""
    base = "https://www.dol.gov/sites/dolgov/files/ETA/Performance/pdfs"
    py = f"PY{program_year}"
    pdf = (
        f"{base}/{py}/PY%20{program_year}%20WIOA%20National%20Performance%20"
        f"Summary.pdf"
    )
    xlsx = f"{base}/{py}/{py}%20Annual%20Report%20Accessible%20File.xlsx"
    return {"pdf": pdf, "xlsx": xlsx}


def _parse_excel(content: bytes) -> tuple[list[dict], dict | None]:
    """Parse the WIOA Annual Report Accessible File.

    Layout: a Performance sheet with one row per (state, program). The state is
    a State Name / State Code pair, the program is in a Program column (WIOA
    Adult, WIOA Dislocated Worker, and so on), and each indicator appears as a
    Target column and an Actual column. We read the Actual columns for the Title
    I Adult program per state, and capture the National (US) row for the vintage
    check, including Dislocated Worker median earnings. Returns (state_rows,
    national); rates stored as fractions (0.74) are converted to percent (74.0).
    """
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = _performance_sheet(wb)
    if ws is None:
        return [], None
    grid = list(ws.iter_rows(values_only=True))
    if not grid:
        return [], None
    header = [("" if c is None else str(c)).strip() for c in grid[0]]
    col = _performance_columns(header)
    if col.get("code") is None or col.get("program") is None:
        return [], None

    rows: list[dict] = []
    national = {"adult": {}, "dislocated_worker": {}}
    for r in grid[1:]:
        program = str(_cell(r, col["program"])).strip()
        code_raw = _cell(r, col["code"])
        name_raw = _cell(r, col.get("state"))
        is_national = (
            str(code_raw).strip().upper() == "US"
            or str(name_raw).strip().lower() == "national"
        )
        adult_vals = {
            "wioa_adult_emp_q2": _rate(r, col.get("emp_q2")),
            "wioa_adult_emp_q4": _rate(r, col.get("emp_q4")),
            "wioa_adult_median_earnings_q2": _money(r, col.get("median")),
            "wioa_adult_credential": _rate(r, col.get("credential")),
            "wioa_adult_msg": _rate(r, col.get("msg")),
        }
        if is_national:
            if program == "WIOA Adult":
                national["adult"] = {
                    "emp_q2": adult_vals["wioa_adult_emp_q2"],
                    "median_earnings_q2": adult_vals["wioa_adult_median_earnings_q2"],
                }
            elif program == "WIOA Dislocated Worker":
                national["dislocated_worker"]["median_earnings_q2"] = _money(
                    r, col.get("median")
                )
            continue
        if program != "WIOA Adult":
            continue
        code = states.try_normalize_code(str(code_raw)) or states.try_normalize_code(
            str(name_raw)
        )
        if not code:
            continue
        record = {"code": code, "name": states.display_name(code)}
        record.update(adult_vals)
        rows.append(record)

    if not (national["adult"] or national["dislocated_worker"]):
        national = None
    return rows, national


def _performance_sheet(wb):
    """Return the per-state performance sheet, or None."""
    for ws in wb.worksheets:
        if ws.title.strip().lower() == "performance":
            return ws
    # Fallback: first sheet whose header names a State Code and a Program column.
    for ws in wb.worksheets:
        first = next(ws.iter_rows(values_only=True), None)
        if not first:
            continue
        low = [("" if c is None else str(c)).strip().lower() for c in first]
        if "state code" in low and "program" in low:
            return ws
    return None


def _performance_columns(header: list[str]) -> dict:
    low = [h.lower() for h in header]

    def find(*needles):
        for i, h in enumerate(low):
            if all(n in h for n in needles):
                return i
        return None

    return {
        "state": find("state name"),
        "code": find("state code"),
        "program": find("program"),
        "emp_q2": find("actual", "employment q2 rate"),
        "emp_q4": find("actual", "employment q4 rate"),
        "median": find("actual", "median earnings"),
        "credential": find("actual", "credential rate"),
        "msg": find("actual", "measurable skills rate"),
    }


def _cell(row, idx):
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return row[idx]


def _rate(row, idx):
    """A WIOA rate stored as a fraction (0.74) becomes a percent (74.0)."""
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None or v == "" or str(v).strip().upper() == "N/A":
        return None
    try:
        num = float(str(v).replace("%", "").replace(",", "").strip())
    except ValueError:
        return None
    if num <= 1.0:
        num *= 100
    return round(num, 1)


def _money(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None or v == "" or str(v).strip().upper() == "N/A":
        return None
    try:
        return round(float(str(v).replace("$", "").replace(",", "").strip()))
    except ValueError:
        return None


def fetch_wioa_py2023(
    program_year: int = 2023, raw_dir: str = "data/raw"
) -> SourceFetch:
    """Fetch WIOA PY2023 state performance from the PDF and Excel files."""
    vintage = f"PY{program_year}"
    urls = _file_urls(program_year)
    attempts: list[dict] = []
    raw_files: list[str] = []
    rows: list[dict] = []
    national: dict | None = None
    blocked = False

    os.makedirs(raw_dir, exist_ok=True)

    # A user-provided Accessible File Excel in data/raw/manual/ takes precedence
    # over the live fetch, so the pipeline can run with no network access.
    manual_xlsx = find_manual(
        raw_dir,
        [
            {"ext": ".xlsx", "contains": ["accessible"]},
            {"ext": ".xlsx", "contains": ["wioa"]},
            {"ext": ".xlsx", "contains": [str(program_year)]},
            {"ext": ".xlsx", "contains": ["annual", "report"]},
        ],
    )
    if manual_xlsx:
        with open(manual_xlsx, "rb") as fh:
            rows, national = _parse_excel(fh.read())
        notes = [f"Parsed user-provided file: {os.path.basename(manual_xlsx)}"]
        if not rows:
            notes.append(
                "No state rows parsed from the provided Excel. Confirm the sheet "
                "names a State column and the Title I indicator columns."
            )
        return SourceFetch(
            source_id=SOURCE_ID,
            requested_vintage=vintage,
            vintage=vintage if rows else None,
            rows=rows,
            national=national,
            provenance={
                "url": "manual",
                "host": "manual_upload",
                "status": None,
                "fetched_at": utc_now_iso(),
                "at_a_glance_not_used": AT_A_GLANCE,
                "raw_files": [manual_xlsx],
                "live_fetch_blocked": False,
                "source_of_record": manual_xlsx,
            },
            blocked=len(rows) == 0,
            notes=notes,
        )

    # Excel first: it carries the state tables and the national total row.
    try:
        xlsx = fetch(urls["xlsx"])
        attempts.append(
            {"url": xlsx.url, "host": xlsx.host, "status": xlsx.status, "blocked": False}
        )
        if xlsx.ok and xlsx.content:
            path = os.path.join(raw_dir, f"wioa_py{program_year}_accessible.xlsx")
            with open(path, "wb") as fh:
                fh.write(xlsx.content)
            raw_files.append(path)
            rows, national = _parse_excel(xlsx.content)
    except FetchBlocked as b:
        attempts.append(
            {"url": b.url, "host": "blocked", "status": None, "blocked": True,
             "deny_reason": b.deny_reason}
        )
        blocked = True

    # PDF second: redundant national summary, parsed only if a backend exists.
    pdf_note = None
    try:
        pdf = fetch(urls["pdf"])
        attempts.append(
            {"url": pdf.url, "host": pdf.host, "status": pdf.status, "blocked": False}
        )
        if pdf.ok and pdf.content:
            path = os.path.join(raw_dir, f"wioa_py{program_year}_summary.pdf")
            with open(path, "wb") as fh:
                fh.write(pdf.content)
            raw_files.append(path)
            if not _pdf_backend_available():
                pdf_note = (
                    "PDF saved but not parsed: no PDF text backend available "
                    "in this environment. State data is read from the Excel."
                )
    except FetchBlocked as b:
        attempts.append(
            {"url": b.url, "host": "blocked", "status": None, "blocked": True,
             "deny_reason": b.deny_reason}
        )
        blocked = True

    notes: list[str] = []
    if pdf_note:
        notes.append(pdf_note)

    if not rows:
        blocked = True
        notes.append(
            "No WIOA PY2023 state rows available in this environment; the DOL "
            "host is blocked by the network policy and the brief carries only "
            "national PY2023 headline figures, not a state table."
        )

    return SourceFetch(
        source_id=SOURCE_ID,
        requested_vintage=vintage,
        vintage=vintage if rows else None,
        rows=rows,
        national=national,
        provenance={
            "url": urls["xlsx"],
            "pdf_url": urls["pdf"],
            "host": "blocked" if blocked and not raw_files else (
                attempts[0]["host"] if attempts else "unknown"
            ),
            "status": None if blocked and not raw_files else (
                attempts[0].get("status") if attempts else None
            ),
            "fetched_at": utc_now_iso(),
            "at_a_glance_not_used": AT_A_GLANCE,
            "archive": ARCHIVE,
            "raw_files": raw_files,
            "attempts": attempts,
            "live_fetch_blocked": blocked,
        },
        blocked=blocked,
        notes=notes,
    )


if __name__ == "__main__":
    out = fetch_wioa_py2023()
    print(f"wioa vintage={out.requested_vintage} rows={len(out.rows)} "
          f"blocked={out.blocked}")
    for note in out.notes:
        print("  note:", note)
